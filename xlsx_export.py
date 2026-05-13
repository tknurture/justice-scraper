"""
Generátor XLSX ze zparsovaných XBRL dat.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="1E3A5F"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }


def _subheader_style():
    return {
        "font": Font(bold=True, size=10),
        "fill": PatternFill("solid", fgColor="D6E4F0"),
        "alignment": Alignment(horizontal="left"),
    }


def _number_format(cell, value):
    if value is None:
        cell.value = ""
        return
    if isinstance(value, float):
        cell.value = value
        cell.number_format = '#,##0'
    else:
        cell.value = value


def _apply_styles(cell, styles: dict):
    for attr, val in styles.items():
        setattr(cell, attr, val)


def _write_statement_sheet(ws, title: str, rows: list[dict], meta: dict):
    """Zapíše jeden list s výkazem."""
    current_date = meta.get("period_current", "Aktuální")
    prior_date = meta.get("period_prior", "Předchozí")

    # Nadpis
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")

    # Záhlaví sloupců
    headers = ["Položka", "Tag (XBRL)", current_date, prior_date]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        for attr, val in _header_style().items():
            setattr(cell, attr, val)

    # Nastavení šířky sloupců
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 20

    # Data
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(rows, start=4):
        label = row.get("label", row.get("tag", ""))
        tag = row.get("tag", "")
        current = row.get("current")
        prior = row.get("prior")

        # Zvýraznění hlavních součtových řádků
        is_total = any(x in label.upper() for x in ["CELKEM", "VÝSLEDEK HOSPODAŘENÍ", "VH"])

        a_cell = ws.cell(row=i, column=1, value=label)
        b_cell = ws.cell(row=i, column=2, value=tag)
        c_cell = ws.cell(row=i, column=3)
        d_cell = ws.cell(row=i, column=4)
        _number_format(c_cell, current)
        _number_format(d_cell, prior)

        if is_total:
            for cell in (a_cell, b_cell, c_cell, d_cell):
                for attr, val in _subheader_style().items():
                    setattr(cell, attr, val)

        for cell in (a_cell, b_cell, c_cell, d_cell):
            cell.border = border
            if isinstance(cell.value, float):
                cell.alignment = Alignment(horizontal="right")

        # Zebra efekt
        if i % 2 == 0 and not is_total:
            for cell in (a_cell, b_cell, c_cell, d_cell):
                cell.fill = PatternFill("solid", fgColor="F7FBFF")

    # Freeze panes
    ws.freeze_panes = "A4"


def build_xlsx(parsed: dict, ico: str, zaverka_info: dict) -> bytes:
    """
    Sestaví XLSX workbook ze zparsovaných dat.
    Vrátí bytes pro stažení.
    """
    wb = Workbook()

    # Sheet 1: Info
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.column_dimensions["A"].width = 30
    ws_info.column_dimensions["B"].width = 50

    info_data = [
        ("IČO", ico),
        ("Období (aktuální)", parsed["meta"].get("period_current", "")),
        ("Období (předchozí)", parsed["meta"].get("period_prior", "")),
        ("Rok závěrky", zaverka_info.get("year", "")),
        ("Zdroj", zaverka_info.get("url", "")),
        ("Počet položek", parsed["meta"].get("total_items", "")),
    ]

    ws_info["A1"].value = "Účetní závěrka - přehled"
    ws_info["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws_info.merge_cells("A1:B1")

    for r, (k, v) in enumerate(info_data, start=3):
        ws_info.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_info.cell(row=r, column=2, value=str(v))

    # Sheet 2: Rozvaha
    if parsed.get("rozvaha"):
        ws_r = wb.create_sheet("Rozvaha")
        _write_statement_sheet(ws_r, "ROZVAHA", parsed["rozvaha"], parsed["meta"])

    # Sheet 3: VZaZ
    if parsed.get("vzaz"):
        ws_v = wb.create_sheet("Výkaz zisku a ztráty")
        _write_statement_sheet(ws_v, "VÝKAZ ZISKU A ZTRÁTY", parsed["vzaz"], parsed["meta"])

    # Sheet 4: Všechna data (raw)
    ws_raw = wb.create_sheet("Všechna data")
    ws_raw.column_dimensions["A"].width = 45
    ws_raw.column_dimensions["B"].width = 20
    ws_raw.column_dimensions["C"].width = 20
    ws_raw.column_dimensions["D"].width = 15

    raw_headers = ["Tag", "Hodnota", "Datum", "Kontext"]
    for col, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        for attr, val in _header_style().items():
            setattr(cell, attr, val)

    for i, item in enumerate(parsed.get("raw", []), start=2):
        ws_raw.cell(row=i, column=1, value=item.get("tag", ""))
        val_cell = ws_raw.cell(row=i, column=2)
        _number_format(val_cell, item.get("value"))
        ws_raw.cell(row=i, column=3, value=item.get("end_date", ""))
        ws_raw.cell(row=i, column=4, value=item.get("context", ""))

    ws_raw.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
